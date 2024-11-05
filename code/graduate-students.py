import pandas as pd
import glob
import os
from openpyxl import load_workbook

# Get all Excel files in the 'data' directory
file_paths = glob.glob("data/*.xlsx")

# Process each file
for file_path in file_paths:
    try:
        # Check if "Graduate Students" sheet exists, and remove it if it does
        with pd.ExcelFile(file_path) as excel_data:
            if "Graduate Students" in excel_data.sheet_names:
                workbook = load_workbook(file_path)
                if "Graduate Students" in workbook.sheetnames:
                    del workbook["Graduate Students"]
                    workbook.save(file_path)
                    workbook.close()
                    print(f"Removed existing 'Graduate Students' sheet from {file_path}")

        # Load the "Full-time Graduate Students" sheet as the base structure
        full_time_df = pd.read_excel(file_path, sheet_name="Full-time Graduate Students")
        part_time_df = pd.read_excel(file_path, sheet_name="Part-time Graduate Students")

        # Ensure "All races" column is excluded from summation
        non_race_columns = [col for col in full_time_df.columns if col != "All races"]

        # Create a copy of the full-time sheet to maintain its structure
        combined_df = full_time_df.copy()

        # Perform element-wise addition on non-race columns, adding values from part-time sheet
        combined_df[non_race_columns] = full_time_df[non_race_columns].add(part_time_df[non_race_columns], fill_value=0)

        # Change the first row label to "All students"
        combined_df.iloc[0, combined_df.columns.get_loc("All races")] = "All students"

        # Write the combined dataframe to a new sheet in the same Excel file
        with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
            combined_df.to_excel(writer, sheet_name="Graduate Students", index=False)
        
        print(f"Processed and saved 'Graduate Students' sheet in {file_path}")

    except Exception as e:
        print(f"Error processing {file_path}: {e}")